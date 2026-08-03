from flask import Flask, render_template, redirect, url_for, request, flash, send_file, jsonify, session
from flask_pymongo import PyMongo
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from bson.objectid import ObjectId
from datetime import datetime, timedelta, date
import pytz
import io
import pandas as pd
import math
import os
import re
import random
import secrets
import json
import requests
import base64
import numpy as np
import cv2

from config import Config

app = Flask(__name__)
app.config.from_object(Config)

if not app.config.get("MONGO_URI"):
    raise RuntimeError("MONGO_URI not set in .env (see README)")

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'info.loginpanel@gmail.com')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', 'wedbfepklgtwtugf')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME', 'info.loginpanel@gmail.com')

GOOGLE_MAPS_API_KEY = os.environ.get('GOOGLE_MAPS_API_KEY', '')
ADMIN_EMAIL = 'admin@jainuniversity.ac.in'

mongo_uri = app.config.get("MONGO_URI", "")
if mongo_uri and "connectTimeoutMS" not in mongo_uri:
    sep = "&" if "?" in mongo_uri else "?"
    app.config["MONGO_URI"] = f"{mongo_uri}{sep}connectTimeoutMS=10000&serverSelectionTimeoutMS=10000"

mongo = None
for attempt in range(5):
    try:
        mongo = PyMongo(app)
        mongo.db.command("ping")
        break
    except Exception as e:
        print(f"[MONGO WARNING] Connection attempt {attempt+1} error: {e}. Retrying in 1s...")
        import time
        time.sleep(1.0)

if not mongo:
    mongo = PyMongo(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"
mail = Mail(app)

# ─────────────────────────────────────────────────────────────
#  PRIMARY DUTY STATION
#  Every "where are you" label in the product comes from here.
# ─────────────────────────────────────────────────────────────
OFFICE_NAME  = "Office of Academics \u00b7 Head Office"
OFFICE_SHORT = "Head Office"
OFFICE_CITY  = "Bengaluru"

OFFICE_LAT = 12.9248224
OFFICE_LNG = 77.5702351
ALLOWED_RADIUS_KM = 0.5       # 500 metres = "at office"
FACE_REQUIRED_RADIUS_KM = 50  # Beyond this, must be in office

IST = pytz.timezone('Asia/Kolkata')
DEFAULT_WORK_HOURS = 8
MIN_WORK_HOURS = 2

SHIFT_TIMINGS = {
    "shift1": {"name": "Shift 1", "start": None, "end": None, "hours": 0},
    "shift2": {"name": "Shift 2", "start": None, "end": None, "hours": 0},
    "normal": {"name": "Normal Login", "start": None, "end": None, "hours": 0}
}

MOTIVATIONAL_QUOTES = [
    "Success is the sum of small efforts repeated day in and day out.",
    "The only way to do great work is to love what you do.",
    "Believe you can and you're halfway there.",
]

UNIVERSITY_HOLIDAYS_2026 = {
   "2026-01-15": {"name": "Uttarayana Punyakala / Makara Sankranti", "type": "general"},
    "2026-01-26": {"name": "Republic Day", "type": "general"},
    "2026-03-19": {"name": "Chandramana Ugadi", "type": "general"},
    "2026-03-21": {"name": "Khutub-E-Ramzan", "type": "general"},
    "2026-03-31": {"name": "Mahaveera Jayanthi", "type": "general"},
    "2026-04-03": {"name": "Good Friday", "type": "general"},
    "2026-04-14": {"name": "Dr. B. R. Ambedkar Jayanthi", "type": "general"},
    "2026-04-20": {"name": "Basava Jayanthi / Akshaya Tritiya", "type": "general"},
    "2026-05-01": {"name": "Labour Day / May Day", "type": "general"},
    "2026-05-28": {"name": "Bakrid (Eid al-Adha)", "type": "general"},
    "2026-06-26": {"name": "Ashura (10th Day of Muharram)", "type": "general"},
    "2026-08-15": {"name": "Independence Day", "type": "general"},
    "2026-08-26": {"name": "Eid-e-Milad", "type": "general"},
    "2026-09-14": {"name": "Swarnagowri Vrata / Varasiddhi Vinayaka Vrata", "type": "general"},
    "2026-10-02": {"name": "Gandhi Jayanthi", "type": "general"},
    "2026-10-10": {"name": "Mahalaya Amavasye", "type": "general"},
    "2026-10-20": {"name": "Maha Navami / Ayudha Pooja", "type": "general"},
    "2026-10-21": {"name": "Vijayadashami", "type": "general"},
    "2026-11-10": {"name": "Balipadyami / Deepavali", "type": "general"},
    "2026-11-27": {"name": "Kanakadasa Jayanthi", "type": "general"},
    "2026-12-25": {"name": "Christmas", "type": "general"},
}


# ─────────────────────────────────────────────────────────────
#  FACE RECOGNITION HELPERS — ROBUST (Multi-photo + Ensemble)
# ─────────────────────────────────────────────────────────────

FACE_CASCADE = cv2.CascadeClassifier(
    cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
)
FACE_CASCADE_ALT = cv2.CascadeClassifier(
    cv2.data.haarcascades + 'haarcascade_frontalface_alt2.xml'
)


def decode_image_b64(b64_data: str):
    """Decode a base64 data-URI or raw base64 to a numpy BGR image."""
    try:
        if ',' in b64_data:
            b64_data = b64_data.split(",", 1)[1]
        b64_data = b64_data.strip()
        padding = 4 - len(b64_data) % 4
        if padding != 4:
            b64_data += '=' * padding
        img_bytes = base64.b64decode(b64_data)
        arr = np.frombuffer(img_bytes, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        return img
    except Exception as e:
        print(f"decode_image_b64 error: {e}")
        return None


def _detect_best_face(gray):
    attempts = [
        (1.05, 5, (50, 50)),
        (1.1,  4, (40, 40)),
        (1.15, 3, (30, 30)),
        (1.2,  2, (25, 25)),
    ]
    for cascade in [FACE_CASCADE, FACE_CASCADE_ALT]:
        if cascade.empty():
            continue
        for scale, neighbors, min_size in attempts:
            try:
                faces = cascade.detectMultiScale(
                    gray, scaleFactor=scale, minNeighbors=neighbors,
                    minSize=min_size, flags=cv2.CASCADE_SCALE_IMAGE
                )
                if len(faces) > 0:
                    return sorted(faces, key=lambda f: f[2] * f[3], reverse=True)[0]
            except Exception:
                continue
    return None


def _center_crop_fallback(img_bgr, size=(100, 100)):
    h, w = img_bgr.shape[:2]
    margin_x = int(w * 0.20)
    margin_y = int(h * 0.10)
    crop = img_bgr[margin_y:h-margin_y, margin_x:w-margin_x]
    if crop.size == 0:
        crop = img_bgr
    return cv2.resize(crop, size)


def _lbp_histogram(gray_img: np.ndarray, grid=(8, 8)) -> np.ndarray:
    """Extract Spatial Grid LBP (Local Binary Pattern) feature signature.
    Divides face into an 8x8 spatial grid (64 sub-regions) for exact spatial feature matching."""
    h, w = gray_img.shape
    gh, gw = h // grid[0], w // grid[1]
    center = gray_img[1:-1, 1:-1].astype(np.int16)
    neighbors = [
        gray_img[0:-2, 0:-2], gray_img[0:-2, 1:-1], gray_img[0:-2, 2:],
        gray_img[1:-1, 2:],   gray_img[2:,   2:],   gray_img[2:,   1:-1],
        gray_img[2:,   0:-2], gray_img[1:-1, 0:-2]
    ]
    lbp = np.zeros_like(center, dtype=np.uint8)
    for bit, nb in enumerate(neighbors):
        lbp |= ((nb.astype(np.int16) >= center).astype(np.uint8) << bit)

    hists = []
    for r in range(grid[0]):
        for c in range(grid[1]):
            sub = lbp[r*gh:(r+1)*gh, c*gw:(c+1)*gw]
            hist, _ = np.histogram(sub.ravel(), bins=256, range=(0, 256), density=True)
            hists.append(hist.astype(np.float32))
    return np.concatenate(hists)


def _extract_face_embedding(img_bgr, required_size=(128, 128), strict=True):
    try:
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        rect = _detect_best_face(gray)
        face_detected = rect is not None

        if rect is not None:
            x, y, w, h = rect
            pad = int(min(w, h) * 0.15)
            h_img, w_img = img_bgr.shape[:2]
            x1 = max(0, x - pad); y1 = max(0, y - pad)
            x2 = min(w_img, x + w + pad); y2 = min(h_img, y + h + pad)
            face_crop_color = img_bgr[y1:y2, x1:x2]
            face_gray_crop  = gray[y1:y2, x1:x2]
        else:
            if strict:
                print("[FACE] Strict face check failed - no face bounding box found")
                return None, None, False
            print("[FACE] Not detected - using center-crop fallback")
            face_crop_color = _center_crop_fallback(img_bgr, size=(200, 200))
            face_gray_crop  = cv2.cvtColor(face_crop_color, cv2.COLOR_BGR2GRAY)

        clahe     = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        face_norm = clahe.apply(face_gray_crop)
        face_rsz  = cv2.resize(face_norm, required_size)
        embedding = _lbp_histogram(face_rsz)
        print(f"[FACE] Extracted spatial embedding, face_detected={face_detected}")
        return embedding.tolist(), face_crop_color, face_detected

    except Exception as e:
        print(f"[FACE] _extract_face_embedding error: {e}")
        return None, None, False


def _chi2_distance(a: np.ndarray, b: np.ndarray) -> float:
    """Calculates region-normalized Chi-Square distance between Spatial LBP signatures."""
    if len(a) != len(b):
        return 2.0
    num = (a - b) ** 2
    den = a + b + 1e-9
    raw_dist = float(np.sum(num / den))
    grid_count = len(a) / 256.0
    return raw_dist / (grid_count if grid_count > 0 else 64.0)


def compare_face_features(stored_list: list, live: list, threshold: float = 0.58):
    live_arr = np.array(live, dtype=np.float32)

    if len(stored_list) > 0 and isinstance(stored_list[0], (int, float)):
        stored_arrays = [np.array(stored_list, dtype=np.float32)]
    else:
        stored_arrays = [np.array(e, dtype=np.float32) for e in stored_list]

    distances = []
    for s in stored_arrays:
        if len(s) == len(live_arr):
            distances.append(_chi2_distance(s, live_arr))
        else:
            # Handle legacy 256-dim embeddings gracefully by recalculating raw chi2
            raw_d = float(np.sum(((s - live_arr[:256])**2) / (s + live_arr[:256] + 1e-9)))
            distances.append(raw_d)

    if not distances:
        return 1.0, False

    min_dist = min(distances)
    return min_dist, min_dist < threshold


def img_to_b64_jpeg(img_bgr, quality: int = 75) -> str:
    _, buf = cv2.imencode(".jpg", img_bgr, [cv2.IMWRITE_JPEG_QUALITY, quality])
    return base64.b64encode(buf).decode("utf-8")


# ─────────────────────────────────────────────────────────────
#  USER MODEL
# ─────────────────────────────────────────────────────────────

class User(UserMixin):
    def __init__(self, user_doc):
        self.id         = str(user_doc["_id"])
        self.username   = user_doc["username"]
        self.role       = user_doc.get("role", "intern")
        self.email      = user_doc.get("email")
        self.work_hours = user_doc.get("work_hours", DEFAULT_WORK_HOURS)


@login_manager.user_loader
def load_user(user_id):
    doc = mongo.db.users.find_one({"_id": ObjectId(user_id)})
    if doc:
        return User(doc)
    return None


# ─────────────────────────────────────────────────────────────
#  UTILITY FUNCTIONS
# ─────────────────────────────────────────────────────────────

def get_ist_now():
    return datetime.now(IST)


def utc_to_ist(utc_dt):
    if utc_dt is None:
        return None
    if isinstance(utc_dt, str):
        try:
            utc_dt = datetime.fromisoformat(utc_dt.replace('Z', '+00:00'))
        except Exception:
            return None
    if utc_dt.tzinfo is None:
        utc_dt = pytz.UTC.localize(utc_dt)
    return utc_dt.astimezone(IST)


def format_ist_time(dt, format_str="%I:%M %p"):
    if dt is None:
        return None
    ist_dt = utc_to_ist(dt)
    if ist_dt is None:
        return None
    return ist_dt.strftime(format_str)


ADDRESS_CACHE = {}

def get_address_from_coords(lat, lng, timeout=1.5):
    if lat is None or lng is None:
        return "Address not available"
    
    try:
        cache_key = (round(float(lat), 4), round(float(lng), 4))
        if cache_key in ADDRESS_CACHE:
            return ADDRESS_CACHE[cache_key]
    except Exception:
        cache_key = None

    if GOOGLE_MAPS_API_KEY:
        try:
            url = f"https://maps.googleapis.com/maps/api/geocode/json?latlng={lat},{lng}&key={GOOGLE_MAPS_API_KEY}"
            response = requests.get(url, timeout=timeout)
            data = response.json()
            if data.get('status') == 'OK' and len(data.get('results', [])) > 0:
                addr = data['results'][0].get('formatted_address', f"Lat: {lat:.6f}, Lng: {lng:.6f}")
                if cache_key: ADDRESS_CACHE[cache_key] = addr
                return addr
        except Exception as e:
            print(f"Google Maps geocode error: {e}")

    # Fallback to OpenStreetMap Nominatim with fast 1.5s timeout
    try:
        headers = {'User-Agent': 'JainAttendanceApp/1.0 (contact@jainuniversity.ac.in)'}
        url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lng}&zoom=18&addressdetails=1"
        response = requests.get(url, headers=headers, timeout=timeout)
        if response.status_code == 200:
            data = response.json()
            display_name = data.get('display_name')
            if display_name:
                if cache_key: ADDRESS_CACHE[cache_key] = display_name
                return display_name
    except Exception as e:
        print(f"Nominatim geocode error: {e}")

    fallback_str = f"Location: {float(lat):.6f}, {float(lng):.6f}"
    
    # Asynchronously resolve in background so API response is instant
    def _async_geocode(l_lat, l_lng, c_key):
        try:
            headers = {'User-Agent': 'JainAttendanceApp/1.0 (contact@jainuniversity.ac.in)'}
            url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={l_lat}&lon={l_lng}&zoom=18&addressdetails=1"
            res = requests.get(url, headers=headers, timeout=5)
            if res.status_code == 200:
                d = res.json()
                d_name = d.get('display_name')
                if d_name:
                    if c_key: ADDRESS_CACHE[c_key] = d_name
                    mongo.db.attendance.update_many(
                        {"login_location.lat": l_lat, "login_location.lng": l_lng},
                        {"$set": {"login_address": d_name, "login_location.address": d_name}}
                    )
        except Exception:
            pass

    import threading
    threading.Thread(target=_async_geocode, args=(lat, lng, cache_key), daemon=True).start()
    return fallback_str



def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi    = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def is_at_office(lat, lng) -> bool:
    if lat is None or lng is None:
        return False
    return haversine_km(lat, lng, OFFICE_LAT, OFFICE_LNG) <= ALLOWED_RADIUS_KM


def calculate_shift_hours(target_hours):
    return target_hours / 2


def _get_lat_lng_from_request():
    lat = lng = None
    if request.is_json:
        j = request.get_json(silent=True) or {}
        lat = j.get("lat")
        lng = j.get("lng")
    if lat is None:
        lat = request.form.get("lat") or request.args.get("lat")
    if lng is None:
        lng = request.form.get("lng") or request.args.get("lng")
    try:
        if lat is not None:
            lat = float(lat)
        if lng is not None:
            lng = float(lng)
    except (ValueError, TypeError):
        lat = lng = None
    return lat, lng


# ─────────────────────────────────────────────────────────────
#  AUTH ROUTES
# ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("user_dashboard") if current_user.role == "intern" else url_for("admin_dashboard"))
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username   = request.form["username"].strip()
        password   = request.form["password"]
        role       = request.form.get("role", "intern")
        email      = request.form.get("email", "")
        work_hours = float(request.form.get("work_hours", DEFAULT_WORK_HOURS))

        if mongo.db.users.find_one({"username": username}):
            flash("Username already exists", "danger")
            return redirect(url_for("register"))

        hashed = bcrypt.generate_password_hash(password).decode("utf-8")
        mongo.db.users.insert_one({
            "username": username, "password": hashed, "role": role,
            "email": email, "work_hours": work_hours,
            "created_at": datetime.utcnow(), "face_registered": False,
            "face_required": False,             # admin must enable this
            "face_registration_enabled": False, # admin must allow registration
            "shift_login_enabled": False        # shifts are off until an admin turns them on
        })
        flash("Registered! Please login.", "success")
        return redirect(url_for("login"))
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        user     = mongo.db.users.find_one({"username": username})
        if not user:
            flash("No such user", "danger")
            return redirect(url_for("login"))
        if bcrypt.check_password_hash(user["password"], password):
            user_obj = User(user)
            login_user(user_obj)
            flash("Logged in", "success")
            return redirect(url_for("user_dashboard") if user_obj.role == "intern" else url_for("admin_dashboard"))
        flash("Invalid credentials", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logged out", "info")
    return redirect(url_for("login"))

@app.route("/admin/impersonate/<user_id>")
@login_required
def impersonate_user(user_id):
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    user_doc = mongo.db.users.find_one({"_id": ObjectId(user_id)})
    if not user_doc:
        flash("User not found", "danger")
        return redirect(url_for("admin_dashboard"))
    
    session["admin_impersonator_id"] = current_user.id
    user_obj = User(user_doc)
    login_user(user_obj)
    flash(f"Impersonating {user_obj.username}", "info")
    return redirect(url_for("user_dashboard"))

@app.route("/admin/revert")
@login_required
def revert_impersonation():
    admin_id = session.get("admin_impersonator_id")
    if not admin_id:
        flash("Not impersonating anyone", "danger")
        return redirect(url_for("login"))
    
    admin_doc = mongo.db.users.find_one({"_id": ObjectId(admin_id)})
    if admin_doc:
        session.pop("admin_impersonator_id", None)
        admin_obj = User(admin_doc)
        login_user(admin_obj)
        flash("Reverted to admin", "success")
        return redirect(url_for("admin_dashboard"))
    
    return redirect(url_for("logout"))


# ─────────────────────────────────────────────────────────────
#  LOCATION CHECK
# ─────────────────────────────────────────────────────────────

@app.route("/api/location/check", methods=["POST"])
@login_required
def check_location():
    data = request.get_json() or {}
    lat  = data.get("lat")
    lng  = data.get("lng")
    try:
        lat = float(lat)
        lng = float(lng)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid coordinates"}), 400

    client_radius = data.get("radius")
    try:
        effective_radius = float(client_radius) if client_radius else ALLOWED_RADIUS_KM
        effective_radius = max(0.05, min(5.0, effective_radius))
    except (TypeError, ValueError):
        effective_radius = ALLOWED_RADIUS_KM

    dist      = haversine_km(lat, lng, OFFICE_LAT, OFFICE_LNG)
    at_office = dist <= effective_radius
    address   = get_address_from_coords(lat, lng)

    # Check if this user has face requirement set by admin
    user_doc = mongo.db.users.find_one({"_id": ObjectId(current_user.id)})
    face_required_by_admin = user_doc.get("face_required", False) if user_doc else False

    return jsonify({
        "at_office":               at_office,
        "distance_km":             round(dist, 3),
        "face_required":           face_required_by_admin,   # now purely admin-driven
        "face_required_by_admin":  face_required_by_admin,
        "address":                 address,
        "office_lat":              OFFICE_LAT,
        "office_lng":              OFFICE_LNG,
        "allowed_radius_km":       effective_radius,
    })


# ─────────────────────────────────────────────────────────────
#  ADMIN — FACE ID CONTROLS
# ─────────────────────────────────────────────────────────────

@app.route("/api/admin/user/<user_id>/face-settings", methods=["POST"])
@login_required
def admin_set_face_settings(user_id):
    """
    Admin endpoint to:
      - Enable/disable face requirement for a user
      - Enable/disable face registration permission for a user
      - Clear face data for a user (force re-registration)
    """
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        data = request.get_json() or {}
        updates = {}

        if "face_required" in data:
            updates["face_required"] = bool(data["face_required"])

        if "face_registration_enabled" in data:
            updates["face_registration_enabled"] = bool(data["face_registration_enabled"])

        if "shift_login_enabled" in data:
            updates["shift_login_enabled"] = bool(data["shift_login_enabled"])

        if data.get("clear_face_data"):
            updates["face_registered"]           = False
            updates["face_features"]             = None
            updates["face_thumb"]                = None
            updates["face_thumbs"]               = None
            updates["face_photo_count"]          = 0
            updates["face_registered_at"]        = None

        if not updates:
            return jsonify({"error": "No settings provided"}), 400

        result = mongo.db.users.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": updates}
        )
        if result.matched_count:
            return jsonify({"ok": True, "updated": updates})
        return jsonify({"error": "User not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/shift-login/bulk", methods=["POST"])
@login_required
def admin_bulk_shift_login():
    """Turn shift attendance on or off for every non-admin account at once."""
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        data = request.get_json() or {}
        if "enabled" not in data:
            return jsonify({"error": "Send an 'enabled' boolean"}), 400
        enabled = bool(data["enabled"])
        result = mongo.db.users.update_many(
            {"role": {"$ne": "admin"}},
            {"$set": {"shift_login_enabled": enabled}}
        )
        return jsonify({
            "ok": True,
            "enabled": enabled,
            "matched": result.matched_count,
            "modified": result.modified_count
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/user/<user_id>/face-settings", methods=["GET"])
@login_required
def admin_get_face_settings(user_id):
    """Get current face settings for a user."""
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        user_doc = mongo.db.users.find_one({"_id": ObjectId(user_id)})
        if not user_doc:
            return jsonify({"error": "User not found"}), 404
        return jsonify({
            "face_required":              user_doc.get("face_required", False),
            "face_registration_enabled":  user_doc.get("face_registration_enabled", False),
            "face_registered":            user_doc.get("face_registered", False),
            "face_photo_count":           user_doc.get("face_photo_count", 0),
            "face_registered_at":         user_doc.get("face_registered_at").isoformat() if user_doc.get("face_registered_at") else None,
            "face_thumb":                 user_doc.get("face_thumb"),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  FACE REGISTRATION  (only allowed when admin has enabled it)
# ─────────────────────────────────────────────────────────────

@app.route("/api/face/register", methods=["POST"])
@login_required
def register_face():
    """
    Accepts a list of up to 5 base64 images.
    NOW REQUIRES admin to have set face_registration_enabled=True for this user.
    """
    try:
        # Check if admin has enabled registration for this user
        user_doc = mongo.db.users.find_one({"_id": ObjectId(current_user.id)})
        if not user_doc:
            return jsonify({"error": "User not found"}), 404

        if not user_doc.get("face_registration_enabled", False):
            return jsonify({
                "error": "Face ID registration has not been enabled for your account. Please contact your administrator.",
                "admin_required": True
            }), 403

        data   = request.get_json() or {}
        images = data.get("images", [])

        if not images:
            single = data.get("image")
            if single:
                images = [single]

        if len(images) < 2:
            return jsonify({"error": "Please capture at least 2 photos for reliable registration (5 recommended)."}), 400
        if len(images) > 5:
            images = images[:5]

        embeddings = []
        face_thumbs = []

        for i, b64 in enumerate(images):
            img = decode_image_b64(b64)
            if img is None:
                return jsonify({"error": f"Could not decode image {i+1}"}), 400
            emb, face_crop, face_detected = _extract_face_embedding(img, strict=True)
            if emb is None or not face_detected:
                return jsonify({"error": f"No clear face detected in photo {i+1}. Please ensure good lighting and look at camera.", "photo_index": i}), 400
            print(f"[REG] Photo {i+1}: face_detected={face_detected}")
            embeddings.append(emb)
            if face_crop is not None:
                face_thumbs.append(img_to_b64_jpeg(face_crop, quality=55))

        CONSISTENCY_THRESHOLD = 1.5
        bad_pairs = []
        all_dists = []
        for i in range(len(embeddings)):
            for j in range(i + 1, len(embeddings)):
                a = np.array(embeddings[i], dtype=np.float32)
                b = np.array(embeddings[j], dtype=np.float32)
                dist = _chi2_distance(a, b)
                all_dists.append(dist)
                if dist >= CONSISTENCY_THRESHOLD:
                    bad_pairs.append((i+1, j+1, round(dist, 3)))

        total_pairs = len(all_dists)
        bad_ratio   = len(bad_pairs) / total_pairs if total_pairs > 0 else 0

        if bad_ratio > 0.70:
            detail = ", ".join([f"photo {p[0]} vs {p[1]}" for p in bad_pairs[:3]])
            return jsonify({
                "error": f"Photos look too different from each other ({detail}). "
                         "Please ensure you are the same person in all photos and use consistent lighting.",
                "bad_pairs": bad_pairs
            }), 400

        mongo.db.users.update_one(
            {"_id": ObjectId(current_user.id)},
            {"$set": {
                "face_features":        embeddings,
                "face_registered":      True,
                "face_registered_at":   datetime.utcnow(),
                "face_photo_count":     len(embeddings),
                "face_thumb":           face_thumbs[0] if face_thumbs else None,
                "face_thumbs":          face_thumbs,
            }}
        )
        return jsonify({
            "ok": True,
            "message": f"Face registered successfully using {len(embeddings)} photos!",
            "photo_count": len(embeddings)
        })
    except Exception as e:
        print(f"Face register error: {e}")
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  FACE VERIFICATION
# ─────────────────────────────────────────────────────────────

@app.route("/api/face/verify", methods=["POST"])
@login_required
def verify_face():
    try:
        data        = request.get_json() or {}
        b64_image   = data.get("image")
        device_info = data.get("device_info", {})
        at_office   = data.get("at_office", False)

        if not b64_image:
            return jsonify({"error": "No image provided"}), 400

        user_doc = mongo.db.users.find_one({"_id": ObjectId(current_user.id)})
        if not user_doc or not user_doc.get("face_registered"):
            return jsonify({
                "error": "Face not registered. Please contact your administrator to enable face registration.",
                "need_registration": True
            }), 400

        stored_features = user_doc.get("face_features")
        if not stored_features:
            return jsonify({"error": "Stored face features missing. Please contact admin to reset face data."}), 400

        img = decode_image_b64(b64_image)
        if img is None:
            return jsonify({"error": "Could not decode image"}), 400

        live_features, face_crop, face_detected = _extract_face_embedding(img, strict=True)
        if live_features is None:
            return jsonify({"error": "No face detected. Please ensure good lighting and look directly at the camera."}), 400

        distance, is_match = compare_face_features(stored_features, live_features)

        ist_now     = get_ist_now()
        snapshot_b64 = img_to_b64_jpeg(img, quality=50)

        snapshot_doc = {
            "user_id":        ObjectId(current_user.id),
            "username":       current_user.username,
            "timestamp":      datetime.utcnow(),
            "timestamp_ist":  ist_now.strftime("%Y-%m-%d %I:%M:%S %p"),
            "date":           ist_now.date().isoformat(),
            "match_result":   is_match,
            "match_distance": round(distance, 4),
            "at_office":      at_office,
            "face_image_b64": snapshot_b64,
            "face_thumb_b64": img_to_b64_jpeg(face_crop, quality=60) if face_crop is not None else None,
            "device_info": {
                "user_agent":           device_info.get("userAgent", ""),
                "platform":             device_info.get("platform", ""),
                "screen":               device_info.get("screen", ""),
                "device_memory":        device_info.get("deviceMemory", ""),
                "hardware_concurrency": device_info.get("cores", ""),
                "timezone":             device_info.get("timezone", ""),
                "imei":                 device_info.get("imei", "Not available (browser)"),
                "device_name":          _extract_device_name(device_info.get("userAgent", "")),
                "browser":              _extract_browser(device_info.get("userAgent", ""))
            },
            "ip_address":       request.remote_addr,
            "reviewed_by_admin": False
        }
        result = mongo.db.face_security_logs.insert_one(snapshot_doc)

        if is_match:
            return jsonify({
                "ok":          True,
                "message":     "Face verified! ✅",
                "snapshot_id": str(result.inserted_id),
                "distance":    round(distance, 4)
            })
        else:
            return jsonify({
                "ok":          False,
                "error":       f"Face mismatch (score {distance:.3f}). Try better lighting or retake from a closer angle.",
                "snapshot_id": str(result.inserted_id),
                "distance":    round(distance, 4)
            }), 401

    except Exception as e:
        print(f"Face verify error: {e}")
        return jsonify({"error": str(e)}), 500


def _extract_device_name(ua: str) -> str:
    ua_lower = ua.lower()
    if "iphone" in ua_lower:  return "iPhone"
    if "ipad"   in ua_lower:  return "iPad"
    if "android" in ua_lower:
        import re
        m = re.search(r"android [0-9.]+; ([^)]+)\)", ua)
        return m.group(1).strip() if m else "Android Device"
    if "windows" in ua_lower: return "Windows PC"
    if "macintosh" in ua_lower or "mac os" in ua_lower: return "Mac"
    if "linux"   in ua_lower: return "Linux Device"
    return "Unknown Device"


def _extract_browser(ua: str) -> str:
    ua_lower = ua.lower()
    if "edg"     in ua_lower: return "Microsoft Edge"
    if "chrome"  in ua_lower: return "Google Chrome"
    if "firefox" in ua_lower: return "Mozilla Firefox"
    if "safari"  in ua_lower: return "Safari"
    if "opera"   in ua_lower or "opr" in ua_lower: return "Opera"
    return "Unknown Browser"


# ─────────────────────────────────────────────────────────────
#  ATTENDANCE ENDPOINTS
# ─────────────────────────────────────────────────────────────

@app.route("/attendance/login", methods=["POST"])
@login_required
def attendance_login():
    if current_user.role != "intern":
        return jsonify({"error": "Only interns can record attendance"}), 403

    data = request.get_json() if request.is_json else {}
    lat, lng = _get_lat_lng_from_request()

    if lat is None or lng is None:
        return jsonify({"error": "Location required"}), 400

    dist      = haversine_km(lat, lng, OFFICE_LAT, OFFICE_LNG)
    at_office = dist <= ALLOWED_RADIUS_KM

    if dist > FACE_REQUIRED_RADIUS_KM:
        address = get_address_from_coords(lat, lng)
        return jsonify({
            "error": "Not within allowed radius",
            "distance_km": round(dist, 4),
            "current_location": {"lat": lat, "lng": lng, "address": address},
            "office_location":  {"lat": OFFICE_LAT, "lng": OFFICE_LNG}
        }), 403

    ist_now   = get_ist_now()
    today     = ist_now.date().isoformat()
    login_type  = data.get("login_type", "normal")
    shift_type  = data.get("shift_type", "normal")
    shift_name  = data.get("shift_name", "Normal Login")
    snapshot_id = data.get("snapshot_id")
    device_info = data.get("device_info", {})

    user             = mongo.db.users.find_one({"_id": ObjectId(current_user.id)})
    work_hours_target = user.get("work_hours", DEFAULT_WORK_HOURS)
    shift_hours       = calculate_shift_hours(work_hours_target)

    if login_type == "shift":
        # Shift attendance is administrator-controlled and off by default.
        # The UI hides it, but the API must refuse it too.
        if not user.get("shift_login_enabled", False):
            return jsonify({
                "error": "Shift attendance is not enabled for your account. "
                         "Ask the Office of Academics to switch it on."
            }), 403

        existing_shift = mongo.db.attendance.find_one({
            "user_id":    ObjectId(current_user.id),
            "date":       today,
            "shift_type": shift_type,
            "logout_time": {"$exists": False}
        })
        if existing_shift:
            return jsonify({"error": f"You already have an active {shift_name} session"}), 400
        session_count  = mongo.db.attendance.count_documents({"user_id": ObjectId(current_user.id), "date": today})
        session_number = session_count + 1
    else:
        existing = mongo.db.attendance.find_one({
            "user_id":    ObjectId(current_user.id),
            "date":       today,
            "logout_time": {"$exists": False}
        })
        if existing:
            return jsonify({"error": "You're already logged in. Please logout first."}), 400
        session_number = 1

    now_utc = datetime.utcnow()
    address = get_address_from_coords(lat, lng)

    attendance_record = {
        "user_id":          ObjectId(current_user.id),
        "username":         current_user.username,
        "date":             today,
        "login_time":       now_utc,
        "login_location":   {"lat": lat, "lng": lng, "address": address},
        "login_address":    address,
        "created_at":       now_utc,
        "login_type":       login_type,
        "shift_type":       shift_type,
        "shift_name":       shift_name,
        "session_number":   session_number,
        "work_hours_target": work_hours_target,
        "shift_target_hours": shift_hours,
        "at_office":        at_office,
        "face_snapshot_id": snapshot_id,
        "face_required":    user.get("face_required", False),
        "device_info": {
            "user_agent":  device_info.get("userAgent", ""),
            "platform":    device_info.get("platform", ""),
            "device_name": device_info.get("deviceName", _extract_device_name(device_info.get("userAgent", ""))),
            "browser":     _extract_browser(device_info.get("userAgent", "")),
            "screen":      device_info.get("screen", ""),
            "imei":        device_info.get("imei", "Not available (browser)"),
            "timezone":    device_info.get("timezone", ""),
            "ip_address":  request.remote_addr
        }
    }
    mongo.db.attendance.insert_one(attendance_record)

    if snapshot_id:
        try:
            mongo.db.face_security_logs.update_one(
                {"_id": ObjectId(snapshot_id)},
                {"$set": {"shift_type": shift_type, "shift_name": shift_name}}
            )
        except Exception as e:
            print(f"Error updating face log shift: {e}")

    # Send email notification
    try:
        user_email = user.get("email")
        if user_email:
            import random
            quotes = [
                "Success is the sum of small efforts, repeated day in and day out.",
                "The only way to do great work is to love what you do.",
                "Don't watch the clock; do what it does. Keep going.",
                "Opportunities don't happen, you create them.",
                "Believe you can and you're halfway there.",
                "Your positive action combined with positive thinking results in success."
            ]
            quote = random.choice(quotes)
            
            logo_path = os.path.join(app.static_folder, 'images', 'applogo.png')
            logo_b64 = ""
            try:
                with open(logo_path, 'rb') as lf:
                    logo_b64 = base64.b64encode(lf.read()).decode('utf-8')
            except Exception:
                pass

            logo_img_tag = f'<img src="data:image/png;base64,{logo_b64}" alt="JAIN Logo" style="height: 48px; width: auto; vertical-align: middle; margin-right: 12px;">' if logo_b64 else ''

            msg = Message(f"Attendance Recorded · {shift_name}", recipients=[user_email])
            msg.html = f"""
            <!DOCTYPE html>
            <html>
            <head>
              <meta charset="utf-8">
              <meta name="viewport" content="width=device-width, initial-scale=1.0">
            </head>
            <body style="margin: 0; padding: 0; background-color: #f4f6f8; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; color: #1c1c1e;">
              <table border="0" cellpadding="0" cellspacing="0" width="100%" style="background-color: #f4f6f8; padding: 30px 15px;">
                <tr>
                  <td align="center">
                    <table border="0" cellpadding="0" cellspacing="0" width="100%" style="max-width: 580px; background-color: #ffffff; border-radius: 16px; overflow: hidden; box-shadow: 0 8px 30px rgba(0,0,0,0.08);">
                      
                      <!-- Header -->
                      <tr>
                        <td style="background-color: #0A1324; padding: 24px 30px;">
                          <table border="0" cellpadding="0" cellspacing="0" width="100%">
                            <tr>
                              <td style="vertical-align: middle; width: 60px;">
                                {logo_img_tag}
                              </td>
                              <td style="vertical-align: middle;">
                                <div style="color: #ffffff; font-size: 15px; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">OFFICE OF ACADEMICS</div>
                                <div style="color: rgba(255,255,255,0.7); font-size: 12px; margin-top: 2px;">JAIN (Deemed-to-be University) · Head Office</div>
                              </td>
                            </tr>
                          </table>
                        </td>
                      </tr>

                      <!-- Body -->
                      <tr>
                        <td style="padding: 32px 30px 24px 30px;">
                          <div style="display: inline-block; background-color: #e8f5e9; color: #2e7d32; font-size: 12px; font-weight: 700; padding: 5px 14px; border-radius: 20px; text-transform: uppercase; margin-bottom: 16px;">
                            ✓ Attendance Recorded
                          </div>
                          
                          <h1 style="font-size: 20px; font-weight: 700; color: #0A1324; margin: 0 0 12px 0;">Hello {current_user.username},</h1>
                          <p style="font-size: 14px; line-height: 1.5; color: #48484a; margin: 0 0 20px 0;">
                            Your attendance sign-in has been successfully registered on the <strong>Attendance Portal</strong>.
                          </p>

                          <!-- Attendance Details Card -->
                          <div style="background-color: #f8f9fa; border: 1px solid #e9ecef; border-radius: 12px; padding: 18px; margin-bottom: 20px;">
                            <table border="0" cellpadding="0" cellspacing="0" width="100%">
                              <tr>
                                <td style="padding: 6px 0; font-size: 13px; color: #6c757d; font-weight: 600;">Signed in User</td>
                                <td style="padding: 6px 0; font-size: 13px; color: #0A1324; font-weight: 700; text-align: right;">{current_user.username}</td>
                              </tr>
                              <tr>
                                <td style="padding: 6px 0; font-size: 13px; color: #6c757d; font-weight: 600; border-top: 1px dashed #e9ecef;">Session / Shift</td>
                                <td style="padding: 6px 0; font-size: 13px; color: #0071e3; font-weight: 700; text-align: right; border-top: 1px dashed #e9ecef;">{shift_name}</td>
                              </tr>
                              <tr>
                                <td style="padding: 6px 0; font-size: 13px; color: #6c757d; font-weight: 600; border-top: 1px dashed #e9ecef;">Sign-In Time (IST)</td>
                                <td style="padding: 6px 0; font-size: 13px; color: #0A1324; font-weight: 700; text-align: right; border-top: 1px dashed #e9ecef;">{format_ist_time(now_utc)}</td>
                              </tr>
                              <tr>
                                <td style="padding: 6px 0; font-size: 13px; color: #6c757d; font-weight: 600; border-top: 1px dashed #e9ecef;">Logged Location</td>
                                <td style="padding: 6px 0; font-size: 12px; color: #212529; font-weight: 600; text-align: right; border-top: 1px dashed #e9ecef; max-width: 220px;">{address}</td>
                              </tr>
                            </table>
                          </div>

                          <!-- Daily Quote -->
                          <div style="border-left: 3px solid #0071e3; background-color: #f0f7ff; padding: 12px 16px; border-radius: 0 8px 8px 0; margin-bottom: 20px;">
                            <p style="font-size: 13px; font-style: italic; color: #1d60a1; margin: 0;">"{quote}"</p>
                          </div>

                          <p style="font-size: 12px; color: #6c757d; margin: 0; line-height: 1.4;">
                            If you did not initiate this sign-in, please inform your team administrator.
                          </p>
                        </td>
                      </tr>

                      <!-- Footer -->
                      <tr>
                        <td style="background-color: #f8f9fa; border-top: 1px solid #e9ecef; padding: 18px 30px; text-align: center;">
                          <div style="font-size: 12px; font-weight: 600; color: #495057;">Office of Academics · JAIN (Deemed-to-be University)</div>
                          <div style="font-size: 11px; color: #868e96; margin-top: 3px;">Head Office, Bengaluru · Attendance Portal</div>
                        </td>
                      </tr>

                    </table>
                  </td>
                </tr>
              </table>
            </body>
            </html>
            """
            mail.send(msg)
    except Exception as e:
        print(f"Failed to send email: {e}")

    return jsonify({
        "ok":               True,
        "login_time":       format_ist_time(now_utc),
        "login_type":       login_type,
        "shift_name":       shift_name,
        "session_number":   session_number,
        "address":          address,
        "shift_target_hours": shift_hours,
        "at_office":        at_office
    })


@app.route("/attendance/logout", methods=["POST"])
@login_required
def attendance_logout():
    if current_user.role != "intern":
        return jsonify({"error": "Only interns can record attendance"}), 403

    data         = request.get_json() if request.is_json else {}
    lat, lng     = _get_lat_lng_from_request()
    force_logout = data.get("force_logout", False)

    ist_now   = get_ist_now()
    today     = ist_now.date().isoformat()
    shift_type = data.get("shift_type")

    query = {"user_id": ObjectId(current_user.id), "date": today, "logout_time": {"$exists": False}}
    if shift_type:
        query["shift_type"] = shift_type

    rec = mongo.db.attendance.find_one(query)
    if not rec:
        rec = mongo.db.attendance.find_one({"user_id": ObjectId(current_user.id), "date": today, "logout_time": {"$exists": False}})
        if not rec:
            return jsonify({"error": "No active login session found"}), 400

    logout_time = datetime.utcnow()
    login_time  = rec.get("login_time")

    if isinstance(login_time, datetime):
        duration = (logout_time - login_time).total_seconds() / 3600.0
    else:
        try:
            parsed   = datetime.fromisoformat(login_time) if isinstance(login_time, str) else None
            duration = (logout_time - parsed).total_seconds() / 3600.0 if parsed else None
        except Exception:
            duration = None

    shift_target_hours = rec.get("shift_target_hours", MIN_WORK_HOURS)

    if not force_logout and duration is not None and duration < MIN_WORK_HOURS:
        return jsonify({
            "warning":      True,
            "message":      f"You've only worked {duration:.1f} hours. Minimum recommended is {MIN_WORK_HOURS} hours. Are you sure?",
            "hours_worked": round(duration, 1),
            "min_hours":    MIN_WORK_HOURS
        }), 200

    updates = {"logout_time": logout_time, "hours": round(duration, 4) if duration is not None else None}
    if lat is not None and lng is not None:
        address = get_address_from_coords(lat, lng)
        updates["logout_location"] = {"lat": lat, "lng": lng, "address": address}
        updates["logout_address"]  = address

    mongo.db.attendance.update_one({"_id": rec["_id"]}, {"$set": updates})

    total_daily_hours = shift1_hours = shift2_hours = 0
    for session in mongo.db.attendance.find({"user_id": ObjectId(current_user.id), "date": today}):
        h = session.get("hours", 0) or 0
        total_daily_hours += h
        if session.get("shift_type") == "shift1": shift1_hours += h
        elif session.get("shift_type") == "shift2": shift2_hours += h

    target_hours   = rec.get("work_hours_target", DEFAULT_WORK_HOURS)
    target_achieved = total_daily_hours >= target_hours
    shift_target   = target_hours / 2

    return jsonify({
        "ok":               True,
        "logout_time":      format_ist_time(logout_time),
        "hours":            updates["hours"],
        "total_daily_hours": round(total_daily_hours, 1),
        "target_hours":     target_hours,
        "target_achieved":  target_achieved,
        "shift1_hours":     round(shift1_hours, 1),
        "shift2_hours":     round(shift2_hours, 1),
        "shift1_completed": shift1_hours >= shift_target,
        "shift2_completed": shift2_hours >= shift_target,
        "shift_target":     round(shift_target, 1),
        "can_login_shift1": not (shift1_hours >= shift_target),
        "can_login_shift2": not (shift2_hours >= shift_target),
    })


# ─────────────────────────────────────────────────────────────
#  ADMIN – FACE SECURITY LOGS
# ─────────────────────────────────────────────────────────────

@app.route("/api/admin/face-logs")
@login_required
def get_face_logs():
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        page        = int(request.args.get("page", 1))
        per_page    = int(request.args.get("per_page", 20))
        user_filter = request.args.get("username")
        match_filter = request.args.get("match")
        date_filter  = request.args.get("date")
        shift_filter = request.args.get("shift")
        query = {}
        if user_filter:              query["username"]     = user_filter
        if match_filter is not None: query["match_result"] = (match_filter.lower() == "true")
        if date_filter:              query["date"]         = date_filter
        if shift_filter:             query["shift_type"]   = shift_filter
        total = mongo.db.face_security_logs.count_documents(query)
        logs  = list(
            mongo.db.face_security_logs.find(query, {"face_image_b64": 0})
            .sort("timestamp", -1).skip((page - 1) * per_page).limit(per_page)
        )
        result = []
        for log in logs:
            result.append({
                "id":               str(log["_id"]),
                "username":         log.get("username"),
                "timestamp_ist":    log.get("timestamp_ist"),
                "date":             log.get("date"),
                "match_result":     log.get("match_result"),
                "match_distance":   log.get("match_distance"),
                "at_office":        log.get("at_office"),
                "face_thumb_b64":   log.get("face_thumb_b64"),
                "shift_type":       log.get("shift_type", "normal"),
                "shift_name":       log.get("shift_name", "Normal Login"),
                "device_info":      log.get("device_info", {}),
                "ip_address":       log.get("ip_address"),
                "reviewed_by_admin": log.get("reviewed_by_admin", False)
            })
        return jsonify({"logs": result, "total": total, "page": page, "per_page": per_page})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/face-logs/<log_id>/image")
@login_required
def get_face_log_image(log_id):
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        log = mongo.db.face_security_logs.find_one({"_id": ObjectId(log_id)})
        if not log:
            return jsonify({"error": "Log not found"}), 404
        mongo.db.face_security_logs.update_one({"_id": ObjectId(log_id)}, {"$set": {"reviewed_by_admin": True}})
        return jsonify({
            "face_image_b64": log.get("face_image_b64"),
            "username":       log.get("username"),
            "timestamp_ist":  log.get("timestamp_ist"),
            "device_info":    log.get("device_info", {}),
            "match_result":   log.get("match_result"),
            "at_office":      log.get("at_office"),
            "lat":            log.get("lat") or log.get("device_info", {}).get("lat"),
            "lng":            log.get("lng") or log.get("device_info", {}).get("lng"),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/face-logs/<log_id>/delete", methods=["DELETE", "POST"])
@login_required
def delete_face_log(log_id):
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        result = mongo.db.face_security_logs.delete_one({"_id": ObjectId(log_id)})
        if result.deleted_count:
            return jsonify({"ok": True, "message": "Face log deleted."})
        return jsonify({"error": "Log not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/face-logs/purge-old", methods=["POST"])
@login_required
def purge_old_face_logs():
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        days    = int((request.get_json() or {}).get("days", 30))
        cutoff  = datetime.utcnow() - timedelta(days=days)
        result  = mongo.db.face_security_logs.delete_many({"timestamp": {"$lt": cutoff}})
        return jsonify({"ok": True, "deleted": result.deleted_count,
                        "message": f"Deleted {result.deleted_count} logs older than {days} days."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/face/status")
@login_required
def face_status():
    user_doc = mongo.db.users.find_one({"_id": ObjectId(current_user.id)})
    return jsonify({
        "face_registered":            bool(user_doc and user_doc.get("face_registered")),
        "face_required":              bool(user_doc and user_doc.get("face_required", False)),
        "face_registration_enabled":  bool(user_doc and user_doc.get("face_registration_enabled", False)),
        "photo_count":                user_doc.get("face_photo_count", 0) if user_doc else 0,
        "registered_at":              user_doc.get("face_registered_at").isoformat() if user_doc and user_doc.get("face_registered_at") else None
    })


# ─────────────────────────────────────────────────────────────
#  ADMIN – USER STATS
# ─────────────────────────────────────────────────────────────

@app.route("/api/admin/user-stats/<user_id>")
@login_required
def admin_user_stats(user_id):
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        user_doc = mongo.db.users.find_one({"_id": ObjectId(user_id)})
        if not user_doc:
            return jsonify({"error": "User not found"}), 404

        cutoff = (date.today() - timedelta(days=365)).isoformat()
        records = list(mongo.db.attendance.find(
            {"user_id": ObjectId(user_id), "date": {"$gte": cutoff}}
        ))
        today_str = date.today().isoformat()
        this_month = date.today().strftime("%Y-%m")

        total_working_days = len(set(r["date"] for r in records))
        attendance_this_month = len(set(r["date"] for r in records if r.get("date", "").startswith(this_month)))
        leaves_this_month = mongo.db.leave_applications.count_documents({
            "user_id": ObjectId(user_id),
            "date": {"$regex": f"^{this_month}"},
            "status": "approved"
        })

        shift_usage = {}
        for rec in records:
            st = rec.get("shift_type", "normal")
            sn = rec.get("shift_name", "Normal Login")
            if st not in shift_usage:
                shift_usage[st] = {"name": sn, "count": 0, "total_hours": 0}
            shift_usage[st]["count"] += 1
            shift_usage[st]["total_hours"] += rec.get("hours", 0) or 0

        return jsonify({
            "username":             user_doc.get("username"),
            "email":                user_doc.get("email"),
            "role":                 user_doc.get("role"),
            "work_hours_target":    user_doc.get("work_hours", DEFAULT_WORK_HOURS),
            "face_registered":      user_doc.get("face_registered", False),
            "face_required":        user_doc.get("face_required", False),
            "face_registration_enabled": user_doc.get("face_registration_enabled", False),
            "face_photo_count":     user_doc.get("face_photo_count", 0),
            "face_thumb":           user_doc.get("face_thumb"),
            "total_working_days":   total_working_days,
            "attendance_this_month": attendance_this_month,
            "leaves_this_month":    leaves_this_month,
            "shift_usage":          shift_usage,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  DASHBOARD DATA
# ─────────────────────────────────────────────────────────────

@app.route("/api/dashboard-data")
@login_required
def get_dashboard_data():
    if current_user.role != "intern":
        return jsonify({"error": "Only interns can access this dashboard"}), 403
    try:
        ist_now = get_ist_now()
        today   = ist_now.date().isoformat()

        today_sessions = list(mongo.db.attendance.find(
            {"user_id": ObjectId(current_user.id), "date": today}
        ).sort("login_time", 1))

        today_data          = []
        total_hours_today   = 0
        shift1_hours = shift2_hours = 0
        active_session      = None

        now_utc = datetime.utcnow()
        for session in today_sessions:
            login_time  = session.get("login_time")
            logout_time = session.get("logout_time")
            
            if logout_time is None and login_time:
                if isinstance(login_time, datetime):
                    session_hours = max(0.0, (now_utc - login_time).total_seconds() / 3600.0)
                else:
                    try:
                        parsed_lt = datetime.fromisoformat(str(login_time).replace('Z', '+00:00'))
                        session_hours = max(0.0, (now_utc - parsed_lt.replace(tzinfo=None)).total_seconds() / 3600.0)
                    except Exception:
                        session_hours = session.get("hours", 0) or 0
            else:
                session_hours = session.get("hours", 0) or 0

            total_hours_today += session_hours
            if session.get("shift_type") == "shift1": shift1_hours += session_hours
            elif session.get("shift_type") == "shift2": shift2_hours += session_hours

            login_loc   = session.get("login_location", {}) or {}
            raw_address = login_loc.get("address") or session.get("login_address") or ""
            lat_val     = login_loc.get("lat")
            lng_val     = login_loc.get("lng")

            if not raw_address or raw_address.startswith("Location:") or raw_address.startswith("Lat:"):
                if lat_val is not None and lng_val is not None:
                    raw_address = get_address_from_coords(lat_val, lng_val)
                    mongo.db.attendance.update_one(
                        {"_id": session["_id"]},
                        {"$set": {"login_address": raw_address, "login_location.address": raw_address}}
                    )

            session_info = {
                "session_number": session.get("session_number", 1),
                "login_type":     session.get("login_type", "normal"),
                "shift_type":     session.get("shift_type", "normal"),
                "shift_name":     session.get("shift_name", "Normal Login"),
                "login_time":     format_ist_time(login_time),
                "logout_time":    format_ist_time(logout_time) if logout_time else None,
                "hours":          round(session_hours, 1),
                "is_active":      logout_time is None,
                "login_address":  raw_address or "Address not available",
                "at_office":      session.get("at_office", False),
            }
            if logout_time is None:
                active_session = session_info
            today_data.append(session_info)

        user              = mongo.db.users.find_one({"_id": ObjectId(current_user.id)})
        work_hours_target = user.get("work_hours", DEFAULT_WORK_HOURS)
        shift_target      = work_hours_target / 2
        shift_login_enabled = user.get("shift_login_enabled", False)

        # Single query for history — limit to last 90 days for performance
        cutoff      = (date.today() - timedelta(days=90)).isoformat()
        all_records = list(mongo.db.attendance.find(
            {"user_id": ObjectId(current_user.id), "date": {"$gte": cutoff}},
            {"login_time": 1, "logout_time": 1, "hours": 1, "date": 1,
             "shift_type": 1, "shift_name": 1, "login_type": 1,
             "session_number": 1, "login_location": 1, "logout_location": 1, "at_office": 1}
        ).sort("date", -1).limit(200))

        total_hours = sum([r.get("hours", 0) or 0 for r in all_records])

        all_formatted = []
        for rec in all_records:
            lt  = rec.get("login_time")
            lot = rec.get("logout_time")
            login_loc  = rec.get("login_location", {})
            logout_loc = rec.get("logout_location", {})
            all_formatted.append({
                "date":           rec.get("date", "N/A"),
                "login_time":     format_ist_time(lt) if lt else "N/A",
                "logout_time":    format_ist_time(lot) if lot else "N/A",
                "hours":          str(round(rec.get("hours", 0) or 0, 1)) if rec.get("hours") else "N/A",
                "shift_type":     rec.get("shift_type", "normal"),
                "shift_name":     rec.get("shift_name", "Normal Login"),
                "login_type":     rec.get("login_type", "normal"),
                "session_number": rec.get("session_number", 1),
                "login_address":  login_loc.get("address", "N/A"),
                "logout_address": logout_loc.get("address", "N/A"),
                "login_lat":      login_loc.get("lat"),
                "login_lng":      login_loc.get("lng"),
                "logout_lat":     logout_loc.get("lat"),
                "logout_lng":     logout_loc.get("lng"),
                "at_office":      rec.get("at_office", False),
            })

        leaves      = list(mongo.db.leave_applications.find(
            {"user_id": ObjectId(current_user.id), "status": "approved"},
            {"date": 1}
        ))
        leave_dates = [l["date"] for l in leaves]

        return jsonify({
            "username":               current_user.username,
            "face_registered":        bool(user.get("face_registered")),
            "face_required":          bool(user.get("face_required", False)),
            "face_registration_enabled": bool(user.get("face_registration_enabled", False)),
            "face_photo_count":       user.get("face_photo_count", 0),
            "today_sessions":         today_data,
            "active_session":         active_session,
            "total_hours_today":      round(total_hours_today, 1),
            "shift1_hours":           round(shift1_hours, 1),
            "shift2_hours":           round(shift2_hours, 1),
            "work_hours_target":      work_hours_target,
            "shift_target":           round(shift_target, 1),
            "shift1_completed":       shift1_hours >= shift_target,
            "shift2_completed":       shift2_hours >= shift_target,
            "can_login_shift1":       not (shift1_hours >= shift_target),
            "can_login_shift2":       not (shift2_hours >= shift_target),
            "total_hours_yeartodate": f"{total_hours:.1f}",
            "office_lat":             OFFICE_LAT,
            "office_lng":             OFFICE_LNG,
            "allowed_radius_km":      ALLOWED_RADIUS_KM,
            "face_required_radius_km": FACE_REQUIRED_RADIUS_KM,
            "office_name":             OFFICE_NAME,
            "office_short":            OFFICE_SHORT,
            "office_city":             OFFICE_CITY,
            "shift_login_enabled":     shift_login_enabled,
            "history":                all_formatted,
            "leaves":                 leave_dates,
            "shifts":                 SHIFT_TIMINGS,
            "min_hours_warning":      MIN_WORK_HOURS
        })
    except Exception as e:
        print(f"Error in dashboard data API: {e}")
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  REMAINING ROUTES (statistics, leave, holidays, admin, export)
# ─────────────────────────────────────────────────────────────

@app.route("/api/statistics", methods=["POST"])
@login_required
def get_statistics():
    if current_user.role != "intern":
        return jsonify({"error": "Only interns can access statistics"}), 403
    try:
        data       = request.get_json()
        start_date = data.get("start_date")
        end_date   = data.get("end_date")
        if not start_date or not end_date:
            return jsonify({"error": "Start and end dates required"}), 400
        sessions   = list(mongo.db.attendance.find({"user_id": ObjectId(current_user.id), "date": {"$gte": start_date, "$lte": end_date}}).sort("date", 1))
        total_days = len(set(s.get("date") for s in sessions))
        total_hours = sum(s.get("hours", 0) or 0 for s in sessions)
        avg_hours_per_day = total_hours / total_days if total_days > 0 else 0
        shift1_hours = sum(s.get("hours", 0) or 0 for s in sessions if s.get("shift_type") == "shift1")
        shift2_hours = sum(s.get("hours", 0) or 0 for s in sessions if s.get("shift_type") == "shift2")
        normal_hours = sum(s.get("hours", 0) or 0 for s in sessions if s.get("login_type") == "normal")
        early_logouts = [s for s in sessions if s.get("hours") and s.get("work_hours_target") and s.get("hours") < s.get("work_hours_target")]
        leaves = list(mongo.db.leave_applications.find({"user_id": ObjectId(current_user.id), "date": {"$gte": start_date, "$lte": end_date}, "status": "approved"}))
        return jsonify({
            "ok": True, "period": {"start": start_date, "end": end_date},
            "total_days": total_days, "total_hours": round(total_hours, 1),
            "avg_hours_per_day": round(avg_hours_per_day, 1),
            "shift1_hours": round(shift1_hours, 1), "shift2_hours": round(shift2_hours, 1),
            "normal_hours": round(normal_hours, 1),
            "shift1_sessions": len([s for s in sessions if s.get("shift_type") == "shift1"]),
            "shift2_sessions": len([s for s in sessions if s.get("shift_type") == "shift2"]),
            "normal_sessions": len([s for s in sessions if s.get("login_type") == "normal"]),
            "early_logouts": len(early_logouts),
            "leaves_taken": len(leaves),
            "leave_details": [{"date": l["date"], "type": l["type"]} for l in leaves]
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/leave/apply", methods=["POST"])
@login_required
def apply_leave():
    if current_user.role != "intern":
        return jsonify({"error": "Only interns can apply for leave"}), 403
    try:
        if not request.is_json:
            return jsonify({"error": "Content-Type must be application/json"}), 415
        data        = request.get_json()
        leave_date  = data.get("date")
        leave_type  = data.get("type")
        comments    = data.get("comments", "")
        if not leave_date or not leave_type:
            return jsonify({"error": "Date and type are required"}), 400
        try:
            parsed_date = datetime.strptime(leave_date, "%Y-%m-%d")
            if parsed_date.date() < date.today():
                return jsonify({"error": "Cannot apply for leave in the past"}), 400
        except ValueError:
            return jsonify({"error": "Invalid date format"}), 400
        existing = mongo.db.leave_applications.find_one({"user_id": ObjectId(current_user.id), "date": leave_date})
        if existing:
            if existing.get("status") in ("pending", "approved"):
                return jsonify({"error": f"You already have a {existing['status']} leave for this date"}), 400
            mongo.db.leave_applications.delete_one({"_id": existing["_id"]})
        if leave_date in UNIVERSITY_HOLIDAYS_2026:
            return jsonify({"error": f"Already a holiday: {UNIVERSITY_HOLIDAYS_2026[leave_date]['name']}"}), 400
        if datetime.strptime(leave_date, "%Y-%m-%d").weekday() in [6]:
            return jsonify({"error": "Cannot apply for leave on Sundays"}), 400
        result = mongo.db.leave_applications.insert_one({
            "user_id": ObjectId(current_user.id), "username": current_user.username,
            "user_email": current_user.email or "", "date": leave_date,
            "type": leave_type, "comments": comments, "status": "pending",
            "applied_at": datetime.utcnow(), "updated_at": datetime.utcnow(), "user_notified": False
        })
        return jsonify({"ok": True, "message": "Leave application submitted", "leave_id": str(result.inserted_id)})
    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500


@app.route("/api/leaves", methods=["GET"])
@login_required
def get_all_leaves():
    if current_user.role != "intern":
        return jsonify({"error": "Only interns can access leave data"}), 403
    try:
        cutoff = (date.today() - timedelta(days=365)).isoformat()
        leaves = list(mongo.db.leave_applications.find({"user_id": ObjectId(current_user.id), "date": {"$gte": cutoff}}))
        return jsonify({l["date"]: {"type": l.get("type"), "comments": l.get("comments",""), "status": l.get("status","pending"), "admin_comments": l.get("admin_comments","")} for l in leaves})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/holidays")
def get_holidays():
    try:
        year  = request.args.get('year', type=int, default=2026)
        month = request.args.get('month', type=int, default=None)
        result = {}
        for date_str, info in UNIVERSITY_HOLIDAYS_2026.items():
            hd = datetime.strptime(date_str, "%Y-%m-%d")
            if hd.year != year: continue
            if month and hd.month != month: continue
            result[date_str] = {"name": info["name"], "type": info["type"], "day": hd.strftime("%A"), "month": hd.strftime("%B"), "date": date_str}
        return jsonify({"holidays": result, "total": len(result), "year": year, "month": month})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/notifications")
@login_required
def get_notifications():
    if current_user.role != "intern":
        return jsonify({"error": "Only interns"}), 403
    try:
        notifications = list(mongo.db.leave_applications.find({"user_id": ObjectId(current_user.id), "status": {"$in": ["approved","denied"]}, "user_notified": {"$ne": True}}).sort("updated_at",-1))
        return jsonify({"notifications": [{"id":str(n["_id"]),"type":"leave_status","date":n.get("date"),"leave_type":n.get("type"),"status":n.get("status"),"admin_comments":n.get("admin_comments",""),"updated_at":n.get("updated_at").isoformat() if n.get("updated_at") else None} for n in notifications]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/notifications/<notif_id>/read", methods=["POST"])
@login_required
def mark_notification_read(notif_id):
    try:
        mongo.db.leave_applications.update_one({"_id": ObjectId(notif_id), "user_id": ObjectId(current_user.id)}, {"$set": {"user_notified": True}})
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/leave-applications")
@login_required
def get_leave_applications():
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        status_filter = request.args.get("status", "all")
        query = {} if status_filter == "all" else {"status": status_filter}
        apps  = list(mongo.db.leave_applications.find(query).sort("applied_at", -1))
        return jsonify({"applications": [{"id":str(a["_id"]),"username":a.get("username"),"user_email":a.get("user_email"),"date":a.get("date"),"type":a.get("type"),"comments":a.get("comments",""),"status":a.get("status","pending"),"admin_comments":a.get("admin_comments",""),"applied_at":a.get("applied_at").isoformat() if a.get("applied_at") else None,"updated_at":a.get("updated_at").isoformat() if a.get("updated_at") else None} for a in apps]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/leave/<leave_id>/update", methods=["POST"])
@login_required
def update_leave_status(leave_id):
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        if not request.is_json:
            return jsonify({"error": "Content-Type must be application/json"}), 415
        data   = request.get_json()
        status = data.get("status")
        admin_comments = data.get("admin_comments","")
        if status not in ["approved","denied"]:
            return jsonify({"error": "Invalid status"}), 400
        leave_app = mongo.db.leave_applications.find_one({"_id": ObjectId(leave_id)})
        if not leave_app:
            return jsonify({"error": "Not found"}), 404
        mongo.db.leave_applications.update_one({"_id": ObjectId(leave_id)}, {"$set": {"status":status,"admin_comments":admin_comments,"updated_at":datetime.utcnow(),"updated_by":current_user.username,"user_notified":False}})
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/update-user-work-hours", methods=["POST"])
@login_required
def update_user_work_hours():
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        data       = request.get_json()
        user_id    = data.get("user_id")
        work_hours = float(data.get("work_hours", DEFAULT_WORK_HOURS))
        if not 1 <= work_hours <= 12:
            return jsonify({"error": "Work hours must be 1–12"}), 400
        result = mongo.db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"work_hours": work_hours}})
        return jsonify({"ok": True, "message": "Work hours updated."}) if result.modified_count else (jsonify({"error": "User not found"}), 404)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/user-locations/<date_str>")
@login_required
def admin_user_locations(date_str):
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        records = list(mongo.db.attendance.find({"date": date_str}).sort("login_time", 1))
        locations = []
        for r in records:
            lt  = r.get("login_time")
            lot = r.get("logout_time")
            login_loc  = r.get("login_location", {})
            logout_loc = r.get("logout_location", {})
            locations.append({
                "username":      r.get("username"),
                "date":          r.get("date"),
                "shift_type":    r.get("shift_type", "normal"),
                "shift_name":    r.get("shift_name", "Normal"),
                "login_time":    format_ist_time(lt) if lt else "N/A",
                "logout_time":   format_ist_time(lot) if lot else None,
                "hours":         str(round(r.get("hours", 0) or 0, 1)),
                "login_address": login_loc.get("address", "N/A"),
                "logout_address": logout_loc.get("address", "N/A"),
                "login_lat":     login_loc.get("lat"),
                "login_lng":     login_loc.get("lng"),
                "logout_lat":    logout_loc.get("lat"),
                "logout_lng":    logout_loc.get("lng"),
                "at_office":     r.get("at_office", False),
                "device_info":   r.get("device_info", {}),
            })
        return jsonify({"locations": locations, "count": len(set(r.get("username") for r in records)), "date": date_str, "office_lat": OFFICE_LAT, "office_lng": OFFICE_LNG})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/user/team-today")
@login_required
def get_team_today():
    """
    Everyone who has signed in today, with the place each of them signed in from.

    NOTE ON VISIBILITY: this endpoint deliberately returns the recorded sign-in
    address for each colleague, not just a coarse "on campus / remote" band.
    That was an explicit product decision. If you ever need to walk it back,
    drop "login_address" from the payload and the UI degrades gracefully.

    Raw GPS coordinates are still withheld for everyone except the caller.
    """
    try:
        today_str = date.today().isoformat()
        # Earliest sign-in first, so the roster reads like an arrival list.
        records = list(mongo.db.attendance.find({"date": today_str}).sort("login_time", 1))

        team_members = []
        seen_users = set()

        for r in records:
            uname = r.get("username")
            if not uname or uname in seen_users:
                continue
            seen_users.add(uname)

            login_loc = r.get("login_location", {}) or {}
            lat = login_loc.get("lat")
            lng = login_loc.get("lng")
            address = login_loc.get("address") or r.get("login_address") or ""

            if not address or address.startswith("Location:") or address.startswith("Lat:"):
                if lat is not None and lng is not None:
                    address = get_address_from_coords(lat, lng)
                    mongo.db.attendance.update_one(
                        {"_id": r["_id"]},
                        {"$set": {"login_address": address, "login_location.address": address}}
                    )

            dist_km = None
            if lat is not None and lng is not None:
                dist_km = round(haversine_km(lat, lng, OFFICE_LAT, OFFICE_LNG), 2)

            at_office = r.get("at_office", False)
            if at_office or (dist_km is not None and dist_km <= ALLOWED_RADIUS_KM):
                loc_status = OFFICE_SHORT
                loc_badge  = "on-campus"
                if not address:
                    address = OFFICE_NAME + ", " + OFFICE_CITY
            elif dist_km is not None and dist_km <= 2.0:
                loc_status = "%s km away" % dist_km
                loc_badge  = "near-campus"
            else:
                loc_status = "Off site"
                loc_badge  = "remote"

            is_self = (uname == current_user.username)
            lt = r.get("login_time")

            team_members.append({
                "username":        uname,
                "is_current_user": is_self,
                "shift_name":      r.get("shift_name", "Normal duty"),
                "login_time":      format_ist_time(lt) if lt else "Active",
                "login_address":   address,
                "location_status": loc_status,
                "location_badge":  loc_badge,
                "dist_km":         dist_km,
                "at_office":       at_office,
                # Exact coordinates stay with the person they belong to.
                "lat": lat if is_self else None,
                "lng": lng if is_self else None,
                "map_lat": lat if is_self else (lat + 0.00025 if lat is not None else None),
                "map_lng": lng if is_self else (lng + 0.00025 if lng is not None else None),
            })

        return jsonify({
            "ok":          True,
            "count":       len(team_members),
            "team":        team_members,
            "office_name": OFFICE_NAME,
            "office_lat":  OFFICE_LAT,
            "office_lng":  OFFICE_LNG,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/early-logouts")
@login_required
def admin_early_logouts():
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        today = date.today().isoformat()
        records = list(mongo.db.attendance.find({"date": today, "logout_time": {"$exists": True}}))
        early = []
        for r in records:
            hours = r.get("hours", 0) or 0
            target = r.get("shift_target_hours") or (r.get("work_hours_target", DEFAULT_WORK_HOURS) / 2)
            if hours < target:
                lt  = r.get("login_time")
                lot = r.get("logout_time")
                early.append({
                    "username":     r.get("username"),
                    "shift_name":   r.get("shift_name", "Normal"),
                    "hours_worked": round(hours, 2),
                    "target_hours": round(target, 2),
                    "shortfall":    round(target - hours, 2),
                    "login_time":   format_ist_time(lt) if lt else "N/A",
                    "logout_time":  format_ist_time(lot) if lot else "N/A",
                })
        return jsonify({"early_logouts": early, "count": len(early)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/analysis-overview")
@login_required
def admin_analysis_overview():
    """Power-BI-style snapshot for one day across EVERY registered (non-admin) user:
    who signed in properly, who left early, who never signed in, who's on leave,
    plus each person's 30-day compliance progress.
    """
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        day = request.args.get("date") or date.today().isoformat()
        try:
            day_date = datetime.strptime(day, "%Y-%m-%d").date()
        except ValueError:
            day_date = date.today()
            day = day_date.isoformat()

        users = list(mongo.db.users.find({"role": {"$ne": "admin"}}))

        # This day's sessions, grouped by user.
        day_recs = list(mongo.db.attendance.find({"date": day}))
        by_uid = {}
        for r in day_recs:
            by_uid.setdefault(str(r.get("user_id")), []).append(r)

        # Approved leaves for the day (match on id or name, whichever the doc carries).
        leaves = list(mongo.db.leave_applications.find({"date": day, "status": "approved"}))
        leave_uids  = {str(l.get("user_id")) for l in leaves if l.get("user_id")}
        leave_names = {l.get("username") for l in leaves if l.get("username")}

        # 30-day window → per-user, per-date total hours, for the progress bars.
        start30 = (day_date - timedelta(days=29)).isoformat()
        recs30  = list(mongo.db.attendance.find({"date": {"$gte": start30, "$lte": day}}))
        prog = {}  # uid -> { date -> total_hours }
        for r in recs30:
            uid = str(r.get("user_id"))
            prog.setdefault(uid, {})
            prog[uid][r.get("date")] = prog[uid].get(r.get("date"), 0.0) + (r.get("hours", 0) or 0)

        buckets = {k: [] for k in ("complete", "active", "early", "absent", "on_leave")}
        totals  = {"registered": 0, "complete": 0, "active": 0, "early": 0, "absent": 0, "on_leave": 0}

        for u in users:
            uid    = str(u["_id"])
            uname  = u.get("username", "")
            target = float(u.get("work_hours", DEFAULT_WORK_HOURS) or DEFAULT_WORK_HOURS)
            recs   = by_uid.get(uid, [])

            login_time = None
            logout_time = None
            total_hours = 0.0
            has_active  = False
            if recs:
                recs_sorted = sorted(recs, key=lambda r: r.get("login_time") or datetime.min)
                login_time  = recs_sorted[0].get("login_time")
                for r in recs:
                    total_hours += r.get("hours", 0) or 0
                    if r.get("logout_time"):
                        logout_time = r.get("logout_time")
                    else:
                        has_active = True

            if not recs:
                status = "on_leave" if (uid in leave_uids or uname in leave_names) else "absent"
            elif has_active:
                status = "active"
            elif total_hours + 1e-6 >= target:
                status = "complete"
            else:
                status = "early"

            # 30-day compliance: share of present days on which the target was met.
            udays = prog.get(uid, {})
            present_days  = len(udays)
            on_track_days = sum(1 for h in udays.values() if h + 1e-6 >= target)
            avg_hours     = round(sum(udays.values()) / present_days, 1) if present_days else 0
            progress_pct  = round(on_track_days / present_days * 100) if present_days else 0

            totals["registered"] += 1
            totals[status]       += 1
            buckets[status].append({
                "user_id":       uid,
                "username":      uname,
                "role":          u.get("role", "intern"),
                "status":        status,
                "target_hours":  round(target, 2),
                "hours":         round(total_hours, 2),
                "shortfall":     round(max(target - total_hours, 0), 2),
                "login_time":    format_ist_time(login_time) if login_time else None,
                "logout_time":   format_ist_time(logout_time) if logout_time else None,
                "present_days":  present_days,
                "on_track_days": on_track_days,
                "avg_hours":     avg_hours,
                "progress_pct":  progress_pct,
            })

        # Sort each bucket most-relevant-first.
        buckets["early"].sort(key=lambda x: x["shortfall"], reverse=True)
        buckets["complete"].sort(key=lambda x: x["progress_pct"], reverse=True)
        for k in ("absent", "on_leave", "active"):
            buckets[k].sort(key=lambda x: x["username"].lower())

        signed_in = totals["complete"] + totals["active"] + totals["early"]
        expected  = totals["registered"] - totals["on_leave"]   # people who were due in
        rate      = round(signed_in / expected * 100) if expected else 0

        return jsonify({
            "date":       day,
            "totals":     totals,
            "signed_in":  signed_in,
            "expected":   expected,
            "rate":       rate,
            "buckets":    buckets,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  PAGE ROUTES
# ─────────────────────────────────────────────────────────────

@app.route("/user/dashboard")
@login_required
def user_dashboard():
    if current_user.role != "intern":
        return redirect(url_for("admin_dashboard"))
    return render_template("user_dashboard.html")


@app.route("/admin/dashboard")
@login_required
def admin_dashboard():
    if current_user.role != "admin":
        return redirect(url_for("user_dashboard"))
    users_raw = list(mongo.db.users.find())
    users     = [{**u, '_id': str(u['_id'])} for u in users_raw]
    recent    = list(mongo.db.attendance.find().sort("date", -1).limit(100))
    cutoff    = (date.today() - timedelta(days=90)).isoformat()
    leaves    = list(mongo.db.leave_applications.find({"date": {"$gte": cutoff}}).sort("applied_at", -1))
    leaves_data = [{"id":str(l["_id"]),"username":l.get("username"),"user_email":l.get("user_email"),"date":l.get("date"),"type":l.get("type"),"comments":l.get("comments",""),"status":l.get("status","pending"),"admin_comments":l.get("admin_comments",""),"applied_at":l.get("applied_at"),"updated_at":l.get("updated_at")} for l in leaves]
    return render_template("admin_dashboard.html", users=users, recent=recent, leaves=leaves_data, shifts=SHIFT_TIMINGS, today_date=date.today().isoformat(), google_maps_api_key=GOOGLE_MAPS_API_KEY)


@app.route("/admin/export", methods=["GET"])
@login_required
def admin_export():
    if current_user.role != "admin":
        return redirect(url_for("admin_dashboard"))
    start = request.args.get("start") or (date.today() - timedelta(days=30)).isoformat()
    end   = request.args.get("end")   or date.today().isoformat()
    recs  = list(mongo.db.attendance.find({"date": {"$gte": start, "$lte": end}}).sort([("date",1),("username",1)]))
    rows  = []
    for r in recs:
        lt  = r.get("login_time")
        lot = r.get("logout_time")
        login_loc  = r.get("login_location", {})
        logout_loc = r.get("logout_location", {})
        di  = r.get("device_info", {})
        rows.append({"username":r.get("username"),"date":r.get("date"),"login_type":r.get("login_type","normal"),"shift_type":r.get("shift_type","normal"),"shift_name":r.get("shift_name","Normal Login"),"session_number":r.get("session_number",1),"login_time":format_ist_time(lt,"%Y-%m-%d %I:%M:%S %p") if lt else "","logout_time":format_ist_time(lot,"%Y-%m-%d %I:%M:%S %p") if lot else "","hours":r.get("hours",0) or 0,"login_lat":login_loc.get("lat"),"login_lng":login_loc.get("lng"),"login_address":login_loc.get("address",""),"logout_lat":logout_loc.get("lat"),"logout_lng":logout_loc.get("lng"),"logout_address":logout_loc.get("address",""),"at_office":r.get("at_office",False),"face_required":r.get("face_required",False),"device_name":di.get("device_name",""),"browser":di.get("browser",""),"ip_address":di.get("ip_address",""),"imei":di.get("imei","")})
    df  = pd.DataFrame(rows)
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    mem = io.BytesIO()
    mem.write(buf.getvalue().encode("utf-8"))
    mem.seek(0)
    return send_file(mem, as_attachment=True, download_name=f"attendance_{start}_to_{end}.csv", mimetype="text/csv")


# ─────────────────────────────────────────────────────────────
#  SIGN-IN ANALYSIS  —  Excel download + shareable public link
#
#  The "Locations & Map" tab already shows, per day, who signed in,
#  at what time and from where. Admins asked to be able to:
#    • download a single day's sign-in analysis as an Excel workbook,
#    • pick several days, pull them into one report and download that, and
#    • mint a link that opens the very same analysis + Excel download
#      for someone who is NOT logged in (e.g. to forward to HR).
#
#  Shares live in mongo.db.analysis_shares, keyed by an unguessable
#  token. An admin can revoke a share at any time.
# ─────────────────────────────────────────────────────────────

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# (row-dict key, column header) — the shape of one analysis row.
ANALYSIS_COLUMNS = [
    ("date",           "Date"),
    ("username",       "Name"),
    ("shift_name",     "Shift"),
    ("login_time",     "Login time (IST)"),
    ("logout_time",    "Logout time (IST)"),
    ("hours",          "Hours"),
    ("at_office",      "At office"),
    ("login_address",  "Login location"),
    ("login_lat",      "Login lat"),
    ("login_lng",      "Login lng"),
    ("logout_address", "Logout location"),
    ("logout_lat",     "Logout lat"),
    ("logout_lng",     "Logout lng"),
    ("device_name",    "Device"),
    ("browser",        "Browser"),
    ("ip_address",     "IP address"),
]


def parse_dates_arg(raw):
    """A comma/space separated list of ISO dates → sorted, de-duped, validated."""
    if not raw:
        return []
    out = []
    for tok in re.split(r"[,\s]+", str(raw).strip()):
        tok = tok.strip()
        if not tok:
            continue
        try:
            datetime.strptime(tok, "%Y-%m-%d")
        except ValueError:
            continue
        if tok not in out:
            out.append(tok)
    return sorted(out)


def expand_range(frm, to):
    """A From/To pair of ISO dates → every day between them, inclusive."""
    try:
        d0 = datetime.strptime(str(frm).strip(), "%Y-%m-%d").date()
        d1 = datetime.strptime(str(to).strip(), "%Y-%m-%d").date()
    except (ValueError, TypeError, AttributeError):
        return []
    if d0 > d1:
        d0, d1 = d1, d0
    out, d = [], d0
    while d <= d1 and len(out) < 732:   # ~2 years hard cap
        out.append(d.isoformat())
        d += timedelta(days=1)
    return out


def resolve_dates(source):
    """Pull the requested days out of a request: explicit `dates`, else `from`/`to`."""
    getter = source.get
    dates = parse_dates_arg(getter("dates"))
    if not dates:
        dates = expand_range(getter("from"), getter("to"))
    return dates


def parse_users_arg(raw):
    """A list or comma-separated string of usernames → clean list ([] = everyone)."""
    if isinstance(raw, list):
        return [str(u).strip() for u in raw if str(u).strip()]
    if not raw:
        return []
    return [u.strip() for u in str(raw).split(",") if u.strip()]


def slugify_title(title):
    s = re.sub(r"[^A-Za-z0-9._-]+", "-", (title or "").strip()).strip("-_")
    return s[:60] or "signin_analysis"


def build_analysis_rows(dates, usernames=None):
    """One row per attendance session across the given dates (earliest first).

    `usernames`, if given, restricts the result to those people; empty/None
    means everyone.
    """
    if not dates:
        return []
    query = {"date": {"$in": dates}}
    if usernames:
        query["username"] = {"$in": usernames}
    recs = list(mongo.db.attendance.find(query).sort([("date", 1), ("login_time", 1)]))
    rows = []
    for r in recs:
        lt  = r.get("login_time")
        lot = r.get("logout_time")
        li  = r.get("login_location", {})  or {}
        lo  = r.get("logout_location", {}) or {}
        di  = r.get("device_info", {})     or {}
        rows.append({
            "date":           r.get("date", ""),
            "username":       r.get("username", ""),
            "shift_name":     r.get("shift_name", "Normal"),
            "login_time":     format_ist_time(lt, "%Y-%m-%d %I:%M %p") if lt else "",
            "logout_time":    format_ist_time(lot, "%Y-%m-%d %I:%M %p") if lot else "",
            "hours":          round(r.get("hours", 0) or 0, 2),
            "at_office":      "Yes" if r.get("at_office") else "No",
            "login_address":  li.get("address", ""),
            "login_lat":      li.get("lat"),
            "login_lng":      li.get("lng"),
            "logout_address": lo.get("address", ""),
            "logout_lat":     lo.get("lat"),
            "logout_lng":     lo.get("lng"),
            "device_name":    di.get("device_name", ""),
            "browser":        di.get("browser", ""),
            "ip_address":     di.get("ip_address", ""),
        })
    return rows


def build_analysis_workbook(dates, title="Sign-in analysis", usernames=None):
    """Return a BytesIO of a formatted .xlsx of the sign-in analysis for `dates`."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows  = build_analysis_rows(dates, usernames)
    ncols = len(ANALYSIS_COLUMNS)
    navy, gold, paper = "0A1324", "EEBD1B", "F6F4EF"

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    thin = Side(style="thin", color="D8D5CE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title band
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(bold=True, size=15, color="FFFFFF")
    tcell.fill = PatternFill("solid", fgColor=navy)
    tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # Sub-band: which days, and when this was generated
    if dates:
        span = dates[0] if len(dates) == 1 else f"{dates[0]} → {dates[-1]} ({len(dates)} day{'s' if len(dates) != 1 else ''})"
    else:
        span = "No dates selected"
    who = f"{len(usernames)} selected user{'s' if len(usernames) != 1 else ''}" if usernames else "All users"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    scell = ws.cell(row=2, column=1,
                    value=f"{span}   ·   {who}   ·   {len(rows)} sign-in record{'s' if len(rows) != 1 else ''}"
                          f"   ·   generated {format_ist_time(datetime.utcnow(), '%Y-%m-%d %I:%M %p')} IST")
    scell.font = Font(size=10, color="5A5A5A")
    scell.fill = PatternFill("solid", fgColor=paper)
    scell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # Header row
    hrow = 4
    for ci, (_, header) in enumerate(ANALYSIS_COLUMNS, start=1):
        cell = ws.cell(row=hrow, column=ci, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[hrow].height = 22

    # Data rows
    widths = [len(h) for _, h in ANALYSIS_COLUMNS]
    for ri, row in enumerate(rows, start=hrow + 1):
        for ci, (key, _) in enumerate(ANALYSIS_COLUMNS, start=1):
            val = row.get(key)
            if val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left" if isinstance(val, str) else "right")
            if (ri - hrow) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FBFAF7")
            widths[ci - 1] = max(widths[ci - 1], len(str(val)))

    # Column widths (capped) + a golden rule under the header
    from openpyxl.utils import get_column_letter
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 10), 46)
    ws.freeze_panes = f"A{hrow + 1}"

    if not rows:
        ws.merge_cells(start_row=hrow + 1, start_column=1, end_row=hrow + 1, end_column=ncols)
        ecell = ws.cell(row=hrow + 1, column=1, value="No sign-in records for the selected day(s).")
        ecell.alignment = Alignment(horizontal="center")
        ecell.font = Font(italic=True, color="888888")

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return mem


def analysis_download_name(dates, title=None):
    base = slugify_title(title) if title else "signin_analysis"
    if not dates:
        return f"{base}.xlsx"
    if len(dates) == 1:
        return f"{base}_{dates[0]}.xlsx"
    return f"{base}_{dates[0]}_to_{dates[-1]}.xlsx"


@app.route("/admin/analysis/excel")
@login_required
def admin_analysis_excel():
    """Download the sign-in analysis as .xlsx (admin only).

    Accepts a `from`/`to` date range (or an explicit `dates` list), an
    optional `users` filter and an optional `title`.
    """
    if current_user.role != "admin":
        return redirect(url_for("admin_dashboard"))
    dates = resolve_dates(request.args)
    if not dates:
        flash("Pick a From and To date to download.", "danger")
        return redirect(url_for("admin_dashboard"))
    users = parse_users_arg(request.args.get("users"))
    title = (request.args.get("title") or "").strip()[:120] or "Sign-in analysis"
    mem = build_analysis_workbook(dates, title=title, usernames=users)
    return send_file(mem, as_attachment=True,
                     download_name=analysis_download_name(dates, title), mimetype=XLSX_MIME)


@app.route("/api/admin/analysis/share", methods=["POST"])
@login_required
def admin_analysis_create_share():
    """Mint a no-login shareable link for the selected analysis (admin only).

    Accepts `from`/`to` (or `dates`), an optional `users` list and a `title`.
    """
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    body  = request.get_json(silent=True) or {}
    dates = resolve_dates(body)
    if not dates:
        return jsonify({"error": "Pick a From and To date."}), 400
    users = parse_users_arg(body.get("users"))
    title = (body.get("title") or "").strip()[:120] or "Sign-in analysis"
    token = secrets.token_urlsafe(24)
    mongo.db.analysis_shares.insert_one({
        "token":       token,
        "dates":       dates,
        "users":       users,
        "title":       title,
        "created_by":  current_user.username,
        "created_at":  datetime.utcnow(),
        "revoked":     False,
    })
    share_url = url_for("shared_analysis_view", token=token, _external=True)
    return jsonify({"ok": True, "token": token, "url": share_url,
                    "dates": dates, "users": users, "title": title})


@app.route("/api/admin/analysis/shares")
@login_required
def admin_analysis_list_shares():
    """List the live share links (admin only)."""
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    shares = list(mongo.db.analysis_shares.find(
        {"revoked": {"$ne": True}}
    ).sort("created_at", -1).limit(50))
    out = [{
        "token":      s["token"],
        "url":        url_for("shared_analysis_view", token=s["token"], _external=True),
        "title":      s.get("title", "Sign-in analysis"),
        "dates":      s.get("dates", []),
        "users":      s.get("users", []),
        "created_by": s.get("created_by", ""),
        "created_at": format_ist_time(s.get("created_at"), "%Y-%m-%d %I:%M %p") if s.get("created_at") else "",
    } for s in shares]
    return jsonify({"shares": out})


@app.route("/api/admin/analysis/share/<token>/revoke", methods=["POST"])
@login_required
def admin_analysis_revoke_share(token):
    """Turn off a share link so it 404s from then on (admin only)."""
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    res = mongo.db.analysis_shares.update_one(
        {"token": token}, {"$set": {"revoked": True, "revoked_at": datetime.utcnow()}})
    if res.matched_count == 0:
        return jsonify({"error": "Not found"}), 404
    return jsonify({"ok": True})


def _get_live_share(token):
    return mongo.db.analysis_shares.find_one({"token": token, "revoked": {"$ne": True}})


@app.route("/share/analysis/<token>")
def shared_analysis_view(token):
    """Public, no-login read-only view of a shared sign-in analysis."""
    share = _get_live_share(token)
    if not share:
        return render_template("shared_analysis.html", not_found=True, token=token), 404
    dates = share.get("dates", [])
    users = share.get("users", [])
    rows  = build_analysis_rows(dates, users)
    # Group rows by day so the page reads like a diary of arrivals.
    by_day = {}
    for r in rows:
        by_day.setdefault(r["date"], []).append(r)
    days = [{"date": d, "rows": by_day[d]} for d in sorted(by_day.keys())]
    return render_template(
        "shared_analysis.html",
        not_found=False,
        token=token,
        title=share.get("title", "Sign-in analysis"),
        dates=dates,
        users=users,
        days=days,
        total_records=len(rows),
        people=len({r["username"] for r in rows}),
        created_by=share.get("created_by", ""),
        office_lat=OFFICE_LAT,
        office_lng=OFFICE_LNG,
    )


@app.route("/share/analysis/<token>/excel")
def shared_analysis_excel(token):
    """Public, no-login Excel download for a shared analysis."""
    share = _get_live_share(token)
    if not share:
        return ("This share link is no longer available.", 404)
    dates = share.get("dates", [])
    users = share.get("users", [])
    title = share.get("title", "Sign-in analysis")
    mem = build_analysis_workbook(dates, title=title, usernames=users)
    return send_file(mem, as_attachment=True,
                     download_name=analysis_download_name(dates, title), mimetype=XLSX_MIME)


@app.route("/admin/create_user", methods=["POST"])
@login_required
def admin_create_user():
    if current_user.role != "admin":
        return redirect(url_for("user_dashboard"))
    username   = request.form.get("username").strip()
    password   = request.form.get("password")
    role       = request.form.get("role", "intern")
    email      = request.form.get("email", "")
    work_hours = float(request.form.get("work_hours", DEFAULT_WORK_HOURS))
    if mongo.db.users.find_one({"username": username}):
        flash("User exists", "danger"); return redirect(url_for("admin_dashboard"))
    hashed = bcrypt.generate_password_hash(password).decode("utf-8")
    mongo.db.users.insert_one({
        "username": username, "password": hashed, "role": role,
        "email": email, "work_hours": work_hours,
        "created_at": datetime.utcnow(),
        "face_registered": False,
        "face_required": False,
        "face_registration_enabled": False,
        "shift_login_enabled": False        # shifts are off until an admin turns them on
    })
    flash("User created", "success"); return redirect(url_for("admin_dashboard"))


@app.route("/admin/delete_user/<user_id>", methods=["POST"])
@login_required
def admin_delete_user(user_id):
    if current_user.role != "admin":
        return jsonify({"error": "Admin access required"}), 403
    try:
        if str(user_id) == str(current_user.id):
            return jsonify({"error": "Cannot delete own account"}), 400
        user = mongo.db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            return jsonify({"error": "User not found"}), 404
        mongo.db.users.delete_one({"_id": ObjectId(user_id)})
        mongo.db.attendance.delete_many({"user_id": ObjectId(user_id)})
        mongo.db.leave_applications.delete_many({"user_id": ObjectId(user_id)})
        mongo.db.face_security_logs.delete_many({"user_id": ObjectId(user_id)})
        return jsonify({"ok": True, "message": f"User {user['username']} deleted."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/health")
def health():
    return {"status": "ok"}


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)